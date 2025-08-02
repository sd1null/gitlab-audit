import gitlab
import json
import os
import openpyxl
import time
import subprocess
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import email.encoders
import zipfile
import git
import time
import pandas as pd

GITLAB_URL = 'https://gitlab.example.com'
gl = gitlab.Gitlab(url=GITLAB_URL, private_token=str(os.getenv('TOKEN')))

exclusion_list= [] # Список id исключаемых проектов gitlab
id_repo_list = []
commit_data = {}
directory = '/tmp'

COMMASPACE = ', '
e_passwd = os.getenv('PASSWD')

def create_exel_file(filename: str):
    exel_file = openpyxl.Workbook()
    new_sheet = exel_file.active
    headers = ['Repository', 'Commit', 'Commit Date', 'Author', 'Path', 'Matched Line', 'Full String']
    for index, header in enumerate(headers, 1):
        new_sheet.cell(row=1, column=index, value=header)
    exel_file.save(filename=f'/tmp/{filename}.xlsx')

def format_exel_file(filename: str):
    df = pd.read_excel(f'/tmp/{filename}.xlsx')
    df = df.groupby('Commit').head(2)
    df.to_excel(f'/tmp/{filename}.xlsx', index=False)

def sorting_strings(filepath: str):
    with open('exclusion.rules', 'r') as file:
        rules = [line.strip() for line in file]
    if not rules:
        return
    else:
        df = pd.read_excel(filepath)
        df['Full String'] = df['Full String'].astype(str)
        mask = df['Full String'].str.contains('|'.join(rules))
        df = df[~mask]
        df.to_excel(filepath, index=False)

def exclusion_string():
    pass

def get_all_repos():
    projects = gl.projects.list(all=True, archived=True)
    for project in projects:
        if project.id not in exclusion_list:
            id_repo_list.append(project.id)

def get_content_at_line(commit_hash, repo_path, file_path, line_number):
    repo = git.Repo(repo_path)  
    commit = repo.commit(commit_hash)
    try:
        blob = commit.tree[file_path]
        content = blob.data_stream.read().decode('ISO-8859-1').splitlines()
        line_content = content[line_number - 1]
        return line_content
    except IndexError:
        return f"Строка номер {line_number} не найдена в файле {file_path} коммита {commit_hash}"
    except KeyError:
        return

def get_repo_info(project_id: int):
    """Получение url, ssh_url_to_repo, и project_full_path для клона репозитория"""
    project = gl.projects.get(project_id)
    data = project.to_json()
    jdata = json.loads(data)
    return [jdata['path'], jdata['ssh_url_to_repo'], jdata['path_with_namespace']]

def clone_repo(url: str, repo, id):
    repo = git.Repo.clone_from(url,f'/tmp/{repo}-{id}')

def sanitize_value(value):
    # Remove non-ASCII characters and ASCII control characters
    value = re.sub(r'[^\x20-\x7E]+', '', value)  # Removes non-ASCII and control characters
    return value

def create_zip():
    """Создание архива с результатом поиска"""
    try:
        with zipfile.ZipFile(f'/tmp/search.zip', 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.write('/tmp/pd_report.xlsx')
            zf.write('/tmp/g_report.xlsx')
            print('Create search.zip')
    except Exception as e:
        print('Error creating zip -',str(e))

def send_mail(work_time):
    username = 'usermail@mydomen.com' # Почта от которой отправляются отчеты
    password = e_passwd
    to_list = ['devops@mydomen.com','ib@mydomen.com']
    subject = 'Gitlab audit report'
    body = 'Attached search.zip'
    filename = '/tmp/search.zip'
    msg = MIMEMultipart()
    msg['From'] = username
    msg['To'] = COMMASPACE.join(to_list)
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))
    with open(filename, 'rb') as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename=filename.split('/')[-1])
        msg.attach(part)

    server = smtplib.SMTP('smpt.server.mycomp.com', 587) #Домен почтового сервера
    server.starttls()
    server.login(username, password)
    server.sendmail(username, to_list, msg.as_string())
    server.quit()
    print('Email sent successfully.')

def grep_repo(repo, project_full_path, id):
    try:
        fpath = project_full_path.replace('/',':')
        cmd = ["./gitleaks", "detect", "--config", "gitleaks.toml", "-f", "json", "-r", f"/tmp/{fpath}@{repo}@@{id}.json", "-s", f"/tmp/{repo}-{id}"]
        subprocess.check_output(cmd,text=True,stderr=subprocess.STDOUT)
        os.system(f'/bin/rm -rf /tmp/{repo}-{id}')
    except subprocess.CalledProcessError as e:
        return

def create_report():
    for filename in os.listdir(directory):
        if filename.endswith('.json'):
            reponame = (filename.split('.json')[0]).split('@')[1]
            project_full_path = (filename.split('@')[0]).replace(':','/')
            gitlab_id = (filename.split('.json')[0]).split('@@')[1]
            with open(os.path.join(directory, filename), 'r') as file:
                data = json.load(file)
                for item in data:
                    match = item.get('Match', None)
                    author = item.get('Author', None)
                    date = item.get('Date', None)
                    commit = item.get('Commit', None)
                    file = item.get('File', None)
                    tag = item.get('Tags', None)
                    start_line = item.get('StartLine', None)
                    if commit is not None:
                        if commit in commit_data:
                            commit_data[commit].append({
                                'Match': match,
                                'Author': author,
                                'Date': date,
                                'Commit': commit,
                                'File': file,
                                'Repo': reponame,
                                'Path': project_full_path,
                                'Tags': tag,
                                'git_id': gitlab_id,
                                'StartLine': start_line 
                            })
                        else:
                            commit_data[commit] = [{
                                'Match': match,
                                'Author': author,
                                'Date': date,
                                'Commit': commit,
                                'File': file,
                                'Repo': reponame,
                                'Path': project_full_path,
                                'Tags': tag,
                                'git_id': gitlab_id,
                                'StartLine': start_line
                            }]
    df = pd.DataFrame([(commit, block) for commit, blocks in commit_data.items() for block in blocks],columns=['Commit', 'Block'])
    result_df = df.groupby('Commit').head(2)
    commit_dict = result_df.groupby('Commit')['Block'].apply(list).to_dict()
    pd_report = openpyxl.load_workbook(filename='/tmp/pd_report.xlsx')
    g_report = openpyxl.load_workbook(filename='/tmp/g_report.xlsx')
    pdr = pd_report.active
    gr = g_report.active
    for commit, data in commit_dict.items():
        for entry in data:
            tag = entry['Tags']
            if 'pd' in tag:
                try:
                    commit = entry['Commit']
                    match = entry['Match']
                    author = entry['Author']
                    date = entry['Date']
                    file = entry['File']
                    reponame = entry['Repo']
                    path = entry['Path']
                    start_line = entry['StartLine']
                    git_id = entry['git_id']
                    full_path_link = '{}/{}/-/blob/{}/{}'.format(GITLAB_URL, path, commit, file)
                    full_string = get_content_at_line(commit_hash=commit,repo_path=f'/tmp/{reponame}-{git_id}',file_path=file,line_number=start_line)
                    full_string = sanitize_value(full_string)
                    p_dataxls = [reponame, commit, date, author, full_path_link, match, full_string]
                    pdr.append(p_dataxls)
                    p_dataxls.clear()
                except:
                    continue
            else:
                try:
                    commit = entry['Commit']
                    match = entry['Match']
                    author = entry['Author']
                    date = entry['Date']
                    file = entry['File']
                    reponame = entry['Repo']
                    path = entry['Path']
                    start_line = entry['StartLine']
                    git_id = entry['git_id']
                    full_path_link = '{}/{}/-/blob/{}/{}'.format(GITLAB_URL, path, commit, file)
                    full_string = get_content_at_line(commit_hash=commit,repo_path=f'/tmp/{reponame}-{git_id}',file_path=file,line_number=start_line)
                    full_string = sanitize_value(full_string)
                    g_dataxls = [reponame, commit, date, author, full_path_link, match, full_string]
                    gr.append(g_dataxls)
                    g_dataxls.clear()
                except:
                    continue
    pd_report.save(filename='/tmp/pd_report.xlsx')
    g_report.save(filename='/tmp/g_report.xlsx')

def main():
    stime = time.time()
    create_exel_file('pd_report')
    create_exel_file('g_report')
    get_all_repos()
    os.system('mkdir -p /root/.ssh && cp /usr/src/app/id_rsa /root/.ssh/ && cp /usr/src/app/known_hosts /root/.ssh/ && chmod 400 /root/.ssh/id_rsa')
    for id in id_repo_list:
        print(get_repo_info(id))
        result = get_repo_info(id)
        clone_repo(result[1],result[0],id)
        grep_repo(result[0],result[2],id)
    print('Create report...')
    create_report()
    format_exel_file('pd_report')
    format_exel_file('g_report')
    sorting_strings('/tmp/pd_report.xlsx')
    create_zip() 
    etime = time.time()
    wtime = etime - stime
    send_mail(wtime)
    os.system('/bin/rm -rf /tmp/*')
    print(wtime)

if __name__ == "__main__":
    main()
